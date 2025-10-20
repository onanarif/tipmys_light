from __future__ import annotations
from typing import List, Dict, Optional
from datetime import datetime
import io, csv
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from django.conf import settings
from django.contrib.auth.decorators import login_required, user_passes_test
from django.forms import ModelForm, DateTimeInput, Select
from django.http import HttpResponse, JsonResponse, HttpResponseBadRequest
from django.shortcuts import get_object_or_404, render
from django.urls import reverse
from django.utils.dateparse import parse_datetime
from django.utils.timezone import localtime, make_aware, is_naive
from django.views.decorators.http import require_http_methods

from faculty.models import Committee, FacultyProfile
from qbank.models import CourseMaster
from ..models import CourseEvent
from ..forms  import CourseEventForm    

DATETIME_LOCAL_FMT = "%Y-%m-%dT%H:%M"

def _dtlocal(value: Optional[datetime]) -> Optional[str]:
    if not value:
        return None
    if getattr(settings, "USE_TZ", False):
        value = localtime(value)
    return value.strftime(DATETIME_LOCAL_FMT)

def _can_manage(user) -> bool:

    return user.is_superuser

def _fetch_committee_rows(*, committee: Committee, start: Optional[datetime], end: Optional[datetime]) -> List[Dict]:
    qs = (CourseEvent.objects
          .select_related("course__lecturer__user", "course__department", "course__type", "faculty__user")
          .filter(course__committee=committee))

    if start:
        qs = qs.filter(start_date__gte=start)
    if end:
        qs = qs.filter(end_date__lte=end)

    rows: List[Dict] = []
    for ev in qs:
        course = ev.course
        lec = getattr(course, "lecturer", None)
        rows.append({
            "Type": "CourseEvent",
            "EventID": ev.pk,
            "Title": course.name if course else "",
            "Start": ev.start_date,
            "End": ev.end_date,
            "Lecturer": f"{getattr(lec.user, 'first_name', '')} {getattr(lec.user, 'last_name', '')}".strip() if lec else "",
            "Department": getattr(course.department, "name", "") if getattr(course, "department", None) else "",
            "DepartmentDivision": getattr(getattr(course, "department", None), "division", "") or "",
            "CourseType": getattr(getattr(course, "type", None), "name", "") or "",
        })
    return rows

def _export_rows(rows: List[Dict], *, export_format: str, filename: str) -> HttpResponse:
    headers = ["Type","Title","Date","Start","End","Lecturer","Department","DepartmentDivision","CourseType"]

    if export_format == "csv":
        buf = io.StringIO()
        w = csv.writer(buf)
        w.writerow(headers)
        for r in rows:
            w.writerow([
                r.get("Type",""), r.get("Title",""),
                (r.get("Start").strftime("%Y-%m-%d") if r.get("Start") else ""),
                (r.get("Start").strftime("%H:%M") if r.get("Start") else ""),
                (r.get("End").strftime("%H:%M") if r.get("End") else ""),
                r.get("Lecturer",""), r.get("Department",""),
                r.get("DepartmentDivision",""), r.get("CourseType",""),
            ])
        resp = HttpResponse(buf.getvalue(), content_type="text/csv; charset=utf-8")
        resp["Content-Disposition"] = f'attachment; filename="{filename}.csv"'
        return resp

    if export_format == "excel":
        wb = Workbook()
        ws = wb.active
        ws.title = "Events"
        ws.append(headers)
        for r in rows:
            ws.append([
                r.get("Type",""), r.get("Title",""),
                (r.get("Start").strftime("%Y-%m-%d") if r.get("Start") else ""),
                (r.get("Start").strftime("%H:%M") if r.get("Start") else ""),
                (r.get("End").strftime("%H:%M") if r.get("End") else ""),
                r.get("Lecturer",""), r.get("Department",""),
                r.get("DepartmentDivision",""), r.get("CourseType",""),
            ])
        ws.freeze_panes = "A2"
        for i in range(1, len(headers)+1):
            col = get_column_letter(i)
            max_len = max(len(str(c.value)) if c.value else 0 for c in ws[col])
            ws.column_dimensions[col].width = min(max(12, max_len + 2), 50)
        resp = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        resp["Content-Disposition"] = f'attachment; filename="{filename}.xlsx"'
        wb.save(resp)
        return resp

    return HttpResponseBadRequest("Unsupported export format")

@login_required
def committee_detail_eventlist_view(request, committee_id: int):
    committee = get_object_or_404(Committee, pk=committee_id)

    start_str = request.GET.get("start") or ""
    end_str   = request.GET.get("end") or ""
    start = parse_datetime(start_str) if start_str else None
    end   = parse_datetime(end_str) if end_str else None
    if start and end and start > end:
        start, end = end, start

    rows = _fetch_committee_rows(committee=committee, start=start, end=end)

    export_kind = request.GET.get("export")
    if export_kind in {"csv","excel"}:
        fname = f"committee_{committee_id}_events"
        return _export_rows(rows, export_format=export_kind, filename=fname)

    context = {
        "committee": committee,
        "rows": rows,
        "start": start,
        "end": end,
        "can_manage": _can_manage(request.user),
        "title": "Committee — Events",
    }
    return render(request, "faculty/committee_eventlist_light.html", context)

@login_required
@require_http_methods(["GET", "POST"])
def course_event_create(request, committee_id: int):
    committee = get_object_or_404(Committee, pk=committee_id)
    if not _can_manage(request.user):
        return JsonResponse({"ok": False, "error": "Permission denied."}, status=403)

    if request.method == "GET":
        form = CourseEventForm(committee=committee)
        return render(request, "faculty/_event_modal.html", {
            "form": form,
            "title": "Ders Etkinliği Ekle",
            "submit_url": reverse("faculty:course_event_create", args=[committee_id]),
            "method": "POST",
        })

    form = CourseEventForm(request.POST, committee=committee)
    if form.is_valid():
        form.save()
        return JsonResponse({"ok": True})
    html = render(request, "faculty/_event_modal.html", {
        "form": form,
        "title": "Ders Etkinliği Ekle",
        "submit_url": reverse("faculty:course_event_create", args=[committee_id]),
        "method": "POST",
    }).content.decode("utf-8")
    return JsonResponse({"ok": False, "html": html})

@login_required
@require_http_methods(["GET", "POST"])
def course_event_edit(request, pk: int):
    ev = get_object_or_404(CourseEvent, pk=pk)
    if not _can_manage(request.user):
        return JsonResponse({"ok": False, "error": "Permission denied."}, status=403)

    if request.method == "GET":
        form = CourseEventForm(instance=ev, committee=ev.course.committee)
        return render(request, "faculty/_event_modal.html", {
            "form": form,
            "title": "Ders Etkinliği Düzelt",
            "submit_url": reverse("faculty:course_event_edit", args=[pk]),
            "method": "POST",
            "delete_url": reverse("faculty:course_event_delete", args=[pk]),
        })

    form = CourseEventForm(request.POST, instance=ev, committee=ev.course.committee)
    if form.is_valid():
        form.save()
        return JsonResponse({"ok": True})
    html = render(request, "faculty/_event_modal.html", {
        "form": form,
        "title": "Ders Etkinliği Düzelt",
        "submit_url": reverse("faculty:course_event_edit", args=[pk]),
        "method": "POST",
            "delete_url": reverse("faculty:course_event_delete", args=[pk]),
    }).content.decode("utf-8")
    return JsonResponse({"ok": False, "html": html})

@login_required
@require_http_methods(["POST"])
def course_event_delete(request, pk: int):
    ev = get_object_or_404(CourseEvent, pk=pk)
    if not _can_manage(request.user):
        return JsonResponse({"ok": False, "error": "Permission denied."}, status=403)
    ev.delete()
    return JsonResponse({"ok": True})
